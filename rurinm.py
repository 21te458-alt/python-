import os
import re
from tkinter import filedialog, Tk, messagebox
from openpyxl import load_workbook
from datetime import datetime, timedelta

def select_txt_file():
    root = Tk()
    root.withdraw()
    file_path = filedialog.askopenfilename(
        title="社員データのTXTファイルを選択",
        filetypes=[("Text files", "*.txt"), ("All files", "*.*")]
    )
    root.destroy()
    return file_path

def select_excel_file():
    root = Tk()
    root.withdraw()
    file_path = filedialog.askopenfilename(
        title="登録先のExcelファイルを選択（.xlsm）",
        filetypes=[("Excel Macro-Enabled Workbook", "*.xlsm"), ("All files", "*.*")]
    )
    root.destroy()
    return file_path

def get_employee_rate(name, line):
    """根据员工姓名和职位返回时薪、深夜时薪和交通费"""
    
    # 特殊个人的交通费和时薪（根据实际数据）
    special_rates = {
        "王 超": {"hourly": "¥1,300", "night": "¥1,100", "transport": "¥800"},
        "黄 铃": {"hourly": "¥1,250", "night": "¥1,050", "transport": "¥0"},
        "王 榕泽": {"hourly": "¥1,250", "night": "¥1,050", "transport": "¥0"},
        "王 玉强": {"hourly": "¥1,250", "night": "¥1,050", "transport": "¥860"},
        "毕 欢欢": {"hourly": "¥1,200", "night": "¥1,000", "transport": "¥800"},
        "齊藤 鉄也": {"hourly": "¥1,200", "night": "¥1,000", "transport": "¥800"},
        "张 浩东": {"hourly": "¥1,200", "night": "¥1,000", "transport": "¥800"},
        "龙 玲": {"hourly": "¥1,200", "night": "¥1,000", "transport": "¥800"},
        "赵 明": {"hourly": "¥1,300", "night": "¥1,100", "transport": "¥0"},
        "SANJEEV SHRESTHA": {"hourly": "¥1,150", "night": "¥1,000", "transport": "¥800"},
        "杜 林": {"hourly": "¥1,150", "night": "¥1,000", "transport": "¥800"},
        "DEUBA SURESH BAHADUR": {"hourly": "¥1,150", "night": "¥1,000", "transport": "¥800"},
        "UPRETY SURENDRA": {"hourly": "¥1,150", "night": "¥1,000", "transport": "¥800"},
        "THAPA SUJAN": {"hourly": "¥1,150", "night": "¥1,000", "transport": "¥800"},
        "Basneta Pravin": {"hourly": "¥1,150", "night": "¥1,000", "transport": "¥800"},
        "Rai Alok": {"hourly": "¥1,150", "night": "¥1,000", "transport": "¥800"},
        "Prabin Paudel": {"hourly": "¥1,150", "night": "¥1,000", "transport": "¥800"},
        "Khadka Bimarsh": {"hourly": "¥1,150", "night": "¥1,000", "transport": "¥800"},
        "刘 宏涛": {"hourly": "¥1,150", "night": "¥1,000", "transport": "¥800"},
        "Rakesh Bum": {"hourly": "¥1,150", "night": "¥1,000", "transport": "¥800"},
        "Saimon Lama": {"hourly": "¥1,150", "night": "¥1,000", "transport": "¥800"},
        "Bogati Bishnu": {"hourly": "¥1,150", "night": "¥1,000", "transport": "¥800"},
        "Pariyar Rohit": {"hourly": "¥1,150", "night": "¥1,000", "transport": "¥800"},
        "Thapa Amit": {"hourly": "¥1,150", "night": "¥1,000", "transport": "¥800"},
    }
    
    # 检查特殊个人
    for key in special_rates:
        if key in name:
            return special_rates[key]["hourly"], special_rates[key]["night"], special_rates[key]["transport"]
    
    # リーダー
    if "リーダー" in line:
        return "¥1,300", "¥1,100", "¥0"
    
    # サブリーダー
    if "サブリーダー" in line:
        return "¥1,250", "¥1,050", "¥800"
    
    # 女性社員
    if "女" in line:
        return "¥1,200", "¥1,000", "¥800"
    
    # 一般社員
    return "¥1,150", "¥1,000", "¥800"

def parse_txt_info(file_path):
    """从TXT文件中解析日期和员工信息"""
    employees = []
    target_date = ""
    target_weekday = ""
    
    with open(file_path, 'r', encoding='utf-8') as f:
        lines = f.readlines()
    
    current_start = "9:30"
    current_end = "18:30"
    current_break = "1:00"
    
    for i, line in enumerate(lines):
        line = line.strip()
        
        # 第2行：日期
        if i == 1:
            date_match = re.search(r'(\d{4})年(\d+)月(\d+)日', line)
            if date_match:
                year = int(date_match.group(1))
                month = int(date_match.group(2))
                day = int(date_match.group(3))
                target_date = f"{month}月{day}日"
                date_obj = datetime(year, month, day)
                weekday_map = {0: "月", 1: "火", 2: "水", 3: "木", 4: "金", 5: "土", 6: "日"}
                target_weekday = weekday_map[date_obj.weekday()]
        
        # 工作时间行
        time_match = re.search(r'(\d+:\d+)-(\d+:\d+)\s+休憩(\d+)H', line)
        if time_match:
            current_start = time_match.group(1)
            current_end = time_match.group(2)
            break_hours = int(time_match.group(3))
            current_break = f"{break_hours}:00"
        
        # 员工数据行
        emp_match = re.search(r'\d+、\t\[(\d+)\]\s+(.+?)\s*[（(]', line)
        if emp_match:
            code = f"[{emp_match.group(1)}]"
            full_name = emp_match.group(2).strip()
            
            # 出勤时间
            start_time = current_start
            end_time = current_end
            break_time = current_break
            
            # 查找特殊时间
            special_time = re.search(r'[（(](\d{1,2}:\d{2})[）)]', line)
            if special_time:
                start_time = special_time.group(1)
            
            # 获取时薪和交通费
            hourly_wage, night_wage, transport_fee = get_employee_rate(full_name, line)
            
            employees.append({
                "code": code,
                "name": full_name,
                "date": target_date,
                "weekday": target_weekday,
                "start_time": start_time,
                "end_time": end_time,
                "break_time": break_time,
                "hourly_wage": hourly_wage,
                "night_wage": night_wage,
                "break_wage": "¥0",
                "overtime_wage": "¥800",
                "transport_fee": transport_fee,
                "category": "昼間時間帯休憩",
                "work_type": "",
                "work_code": "1"
            })
    
    return employees, target_date, target_weekday

def is_file_open(file_path):
    try:
        with open(file_path, 'r+b') as f:
            return False
    except PermissionError:
        return True
    except:
        return False

def find_last_row_of_previous_date(excel_path, target_date_str):
    """找到目标日期前一天的最后一行的行号"""
    match = re.search(r'(\d+)月(\d+)日', target_date_str)
    if not match:
        return None
    
    month = int(match.group(1))
    day = int(match.group(2))
    target_date_obj = datetime(2026, month, day)
    previous_date_obj = target_date_obj - timedelta(days=1)
    
    print(f"🔍 前日を検索: {previous_date_obj.month}月{previous_date_obj.day}日")
    
    wb = load_workbook(excel_path, keep_vba=True)
    ws = wb.active
    
    previous_date_rows = []
    for row in range(2, ws.max_row + 1):
        cell_j = ws.cell(row=row, column=10)
        cell_value = cell_j.value
        if cell_value and isinstance(cell_value, datetime):
            if cell_value.year == 2026 and cell_value.month == previous_date_obj.month and cell_value.day == previous_date_obj.day:
                previous_date_rows.append(row)
    
    wb.close()
    
    if previous_date_rows:
        last_row = previous_date_rows[-1]
        print(f"✅ 前日の最終行: {last_row}行目")
        return last_row
    else:
        print(f"⚠️ {previous_date_obj.month}月{previous_date_obj.day}日 が見つかりません")
        return None

def write_to_excel(excel_path, employees, start_row):
    """从指定行开始写入数据"""
    if is_file_open(excel_path):
        print("❌ Excelファイルが開いています！閉じてから再実行してください")
        return 0
    
    wb = load_workbook(excel_path, keep_vba=True)
    ws = wb.active
    
    written_count = 0
    for i, emp in enumerate(employees):
        row = start_row + i
        
        try:
            # C列：社員コード + 氏名
            ws.cell(row=row, column=3, value=f"{emp['code']} {emp['name']}")
            # D列：区分
            ws.cell(row=row, column=4, value="◾")
            # E列：作業名
            ws.cell(row=row, column=5, value=emp['category'])
            # F列：請求時給
            ws.cell(row=row, column=6, value=emp['hourly_wage'])
            # G列：給料時給
            ws.cell(row=row, column=7, value=emp['hourly_wage'])
            # H列：請求交通費
            ws.cell(row=row, column=8, value=emp['transport_fee'])
            # I列：給料交通費
            ws.cell(row=row, column=9, value=emp['transport_fee'])
            # J列：日付
            date_match = re.search(r'(\d+)月(\d+)日', emp['date'])
            if date_match:
                month = int(date_match.group(1))
                day = int(date_match.group(2))
                date_obj = datetime(2026, month, day)
                ws.cell(row=row, column=10, value=date_obj)
            # K列：曜日
            ws.cell(row=row, column=11, value=emp['weekday'])
            # L列：日数
            ws.cell(row=row, column=12, value=1)
            # M列：休憩
            ws.cell(row=row, column=13, value=emp['break_time'])
            # N列：開始時間
            ws.cell(row=row, column=14, value=emp['start_time'])
            # O列：終了時間
            ws.cell(row=row, column=15, value=emp['end_time'])
            
            written_count += 1
            print(f"✅ {emp['code']} {emp['name']} | 時給{emp['hourly_wage']} 交通費{emp['transport_fee']} | {emp['start_time']}→{emp['end_time']} → {row}行目")
        except Exception as e:
            print(f"⚠️ {row}行目エラー: {e}")
    
    wb.save(excel_path)
    wb.close()
    return written_count

def main():
    print("=" * 60)
    print("勤怠データ登録ツール")
    print("=" * 60)
    
    print("\n📄 ステップ1: 社員データのTXTファイルを選択してください")
    txt_file = select_txt_file()
    if not txt_file:
        print("❌ キャンセルされました")
        return
    print(f"✅ {os.path.basename(txt_file)}")
    
    print("\n📊 ステップ2: 登録先のExcelファイルを選択してください")
    excel_file = select_excel_file()
    if not excel_file:
        print("❌ キャンセルされました")
        return
    print(f"✅ {os.path.basename(excel_file)}")
    
    if is_file_open(excel_file):
        print("\n❌ エラー: Excelファイルが開いています！")
        messagebox.showerror("エラー", "Excelファイルを閉じてから実行してください")
        return
    
    print("\n🔍 TXTファイルを解析中...")
    employees, target_date, target_weekday = parse_txt_info(txt_file)
    
    if not employees:
        print("❌ 従業員データが見つかりません")
        return
    
    print(f"✅ {len(employees)}件の従業員データを検出")
    print(f"📅 対象日付: {target_date} ({target_weekday})")
    
    print("\n📋 従業員リスト:")
    print("-" * 80)
    for i, emp in enumerate(employees):
        print(f"{i+1:2d}. {emp['code']} {emp['name']:25} | 時給{emp['hourly_wage']} 交通費{emp['transport_fee']} | {emp['start_time']}→{emp['end_time']}")
    
    print(f"\n🔍 {target_date} の前日を検索中...")
    last_row = find_last_row_of_previous_date(excel_file, target_date)
    
    if last_row:
        start_row = last_row + 1
        print(f"\n📝 {start_row}行目から {len(employees)}件のデータを登録します")
    else:
        print("\n⚠️ 前日が見つかりません。手動で開始行を入力してください")
        try:
            start_row = int(input("開始行を入力してください: "))
        except:
            print("❌ キャンセルされました")
            return
    
    confirm = input(f"\n{len(employees)}件のデータを登録しますか？ (y/N): ").strip().lower()
    
    if confirm == 'y' or confirm == 'yes':
        print("\n📝 登録を実行中...")
        count = write_to_excel(excel_file, employees, start_row)
        if count > 0:
            print(f"\n✨ 完了しました！ {count}件のデータを登録しました")
            messagebox.showinfo("完了", f"{count}件のデータを登録しました")
        else:
            print("\n❌ 登録に失敗しました")
    else:
        print("\n❌ キャンセルされました")

if __name__ == "__main__":
    main()