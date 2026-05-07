# -*- coding: utf-8 -*-
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment
import re
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from datetime import datetime
import os

def select_file(title="请选择Excel文件", file_type="all"):
    """打开文件选择对话框"""
    root = tk.Tk()
    root.withdraw()
    root.attributes('-topmost', True)
    
    if file_type == "attendance":
        filetypes = [("Excel文件", "*.xlsx *.xls"), ("所有文件", "*.*")]
    else:
        filetypes = [("启用宏的工作簿", "*.xlsm"), ("Excel文件", "*.xlsx *.xls"), ("所有文件", "*.*")]
    
    file_path = filedialog.askopenfilename(title=title, filetypes=filetypes)
    root.destroy()
    return file_path

def is_english_name(name):
    """判断是否为英文名字（包含英文字母）"""
    return bool(re.search(r'[A-Za-z]', name))

def clean_name_text(name, preserve_spaces=True):
    """清理姓名中的特殊字符"""
    if not name:
        return name
    
    name = str(name).strip()
    
    if preserve_spaces and is_english_name(name):
        name = re.sub(r'[\r\n\t]', '', name)
        name = re.sub(r'\s+', ' ', name)
    else:
        name = re.sub(r'[\s　\n\r\t]', '', name)
    
    return name

def parse_date_to_datetime(date_str, default_year=2026):
    """将各种日期格式转换为datetime对象"""
    if date_str is None:
        return None
    
    if isinstance(date_str, (datetime, pd.Timestamp)):
        return date_str
    
    date_str = str(date_str).strip()
    
    match = re.match(r'(\d{4})-(\d{1,2})-(\d{1,2})', date_str)
    if match:
        return datetime(int(match.group(1)), int(match.group(2)), int(match.group(3)))
    
    match = re.search(r'(\d{4})年(\d{1,2})月(\d{1,2})日', date_str)
    if match:
        return datetime(int(match.group(1)), int(match.group(2)), int(match.group(3)))
    
    match = re.search(r'(\d{1,2})月(\d{1,2})日', date_str)
    if match:
        return datetime(default_year, int(match.group(1)), int(match.group(2)))
    
    return None

def format_date_key(date_obj):
    """格式化日期为月日"""
    if date_obj:
        return f"{date_obj.month}月{date_obj.day}日"
    return None

def read_attendance(file_path):
    """读取出勤表 - 获取所有员工的出勤和加班数据"""
    if not os.path.exists(file_path):
        return None, {}, {}, None, None, {}
    
    df = pd.read_excel(file_path, sheet_name=0, engine='openpyxl')
    df.columns = df.columns.astype(str).str.strip()
    
    # 自动识别日期列
    date_columns = {}
    date_objects = {}
    
    skip_keywords = ['提交时间', '姓名', '加班', '通行证', '签证', 'time', 'name', '备注']
    
    for col in df.columns:
        is_skip = False
        for kw in skip_keywords:
            if kw in col:
                is_skip = True
                break
        if is_skip:
            continue
        
        date_obj = parse_date_to_datetime(col)
        if date_obj:
            key = format_date_key(date_obj)
            date_columns[key] = col
            date_objects[key] = date_obj
    
    # 姓名列
    name_col = None
    for col in df.columns:
        if col == '姓名' or '氏名' in col or '名前' in col:
            name_col = col
            break
    
    if not name_col:
        for col in df.columns:
            if '名' in col:
                name_col = col
                break
    
    if name_col:
        df[name_col] = df[name_col].astype(str).str.strip()
        df[name_col] = df[name_col].apply(lambda x: clean_name_text(x, preserve_spaces=True) if x != 'nan' else x)
    
    # 加班情况统计列
    overtime_col = None
    for col in df.columns:
        if '加班' in col or '残業' in col or 'overtime' in col.lower():
            overtime_col = col
            break
    
    if overtime_col:
        df[overtime_col] = df[overtime_col].astype(str).str.strip()
    
    # 获取每天出勤的人及其加班情况
    attendance_by_date = {}
    for date_key, col_name in date_columns.items():
        df[col_name] = df[col_name].astype(str).str.strip()
        mask = df[col_name] == '出勤'
        attendance_by_date[date_key] = {}
        for idx in df.index[mask]:
            name = df.loc[idx, name_col]
            if name and str(name) != 'nan' and str(name) != 'None' and name != '':
                overtime = df.loc[idx, overtime_col] if overtime_col else None
                clean_name = clean_name_text(name, preserve_spaces=True)
                attendance_by_date[date_key][clean_name] = overtime
    
    return df, date_columns, date_objects, name_col, overtime_col, attendance_by_date

def get_overtime_mark(overtime_value):
    """根据加班情况返回标记"""
    if overtime_value is None:
        return None
    
    overtime_str = str(overtime_value).strip()
    
    if 'D' in overtime_str or '不能加班' in overtime_str:
        return None
    
    if 'A' in overtime_str or 'B' in overtime_str or 'C' in overtime_str:
        return '○'
    
    if '30分钟以内' in overtime_str or '30分钟到1小时' in overtime_str or '1小时以上' in overtime_str:
        return '○'
    
    return None

def is_real_name(value):
    """判断一个值是否为真正的姓名"""
    if value is None:
        return False
    
    val_str = str(value).strip()
    if not val_str:
        return False
    
    if is_english_name(val_str):
        cleaned = re.sub(r'[\r\n\t]', '', val_str)
        cleaned = re.sub(r'\s+', ' ', cleaned).strip()
    else:
        cleaned = re.sub(r'[\s　\n\r\t]', '', val_str)
    
    if not cleaned or len(cleaned) < 2:
        return False
    
    invalid_patterns = [
        r'^[0-9]+$', r'^\d+[、\.]', r'休憩', r'残業', 
        r'9:', r'18:', r'20:', r'関西', r'空港', 
        r'必出勤', r'可出勤', r'^[一二三四五六七八九十]+$',
        r'2026', r'月', r'日', r'9:00', r'18:00'
    ]
    
    for pattern in invalid_patterns:
        if re.search(pattern, cleaned):
            return False
    
    return True

def find_template_staff(ws):
    """从排班表中提取人员名单（第二张表）"""
    staff_list = []
    
    # 找到时间行
    start_row = None
    for row in range(1, min(ws.max_row + 1, 30)):
        for col in range(1, 5):
            cell_val = ws.cell(row, col).value
            if cell_val:
                val_str = str(cell_val).strip()
                if '9:00' in val_str or '休憩' in val_str or '10:00' in val_str or '11:00' in val_str:
                    start_row = row + 1
                    break
        if start_row:
            break
    
    if not start_row:
        start_row = 5
    
    print(f"排班表员工起始行: {start_row}")
    
    # 扫描员工名单（从右侧列，通常15-25列）
    for row in range(start_row, min(start_row + 80, ws.max_row + 1)):
        found_name = None
        for col in range(12, min(ws.max_column + 1, 26)):
            cell_val = ws.cell(row, col).value
            if is_real_name(cell_val):
                name = str(cell_val).strip()
                clean_name = clean_name_text(name, preserve_spaces=True)
                if clean_name and len(clean_name) >= 2:
                    found_name = clean_name
                    break
        
        if found_name:
            if found_name not in staff_list:
                staff_list.append(found_name)
                print(f"  排班表员工: {found_name} (行{row})")
        else:
            if len(staff_list) > 5:
                consecutive_empty = 0
                for r in range(row, min(row + 3, ws.max_row + 1)):
                    has_name = False
                    for col in range(12, 26):
                        if is_real_name(ws.cell(r, col).value):
                            has_name = True
                            break
                    if not has_name:
                        consecutive_empty += 1
                    else:
                        break
                if consecutive_empty >= 2:
                    break
    
    print(f"排班表共找到 {len(staff_list)} 名员工")
    return staff_list, start_row

def find_dates_in_schedule(ws):
    """从排班表中找日期行和列位置"""
    date_info = {}
    
    for row in range(1, min(ws.max_row + 1, 15)):
        for col in range(1, min(ws.max_column + 1, 40)):
            cell_val = ws.cell(row, col).value
            if cell_val:
                date_obj = parse_date_to_datetime(str(cell_val))
                if date_obj:
                    date_key = format_date_key(date_obj)
                    date_info[date_key] = (col, col + 1)
                    print(f"  发现日期: {date_key} 在列 {col}")
    
    return date_info

def fill_schedule(attendance_by_date, att_date_map, att_date_objects, template_path, output_path):
    """填写可出勤人员表 - 合并排班表名单和出勤表出勤人员"""
    wb = load_workbook(template_path, keep_vba=True)
    ws = wb.active
    
    # 获取排班表的人员名单
    template_staff, start_row = find_template_staff(ws)
    
    if len(template_staff) == 0:
        print("错误：排班表中未找到人员名单！")
        messagebox.showerror("错误", "排班表中未找到人员名单！")
        return 0
    
    print(f"\n排班表人员名单: {template_staff}")
    
    # 获取排班表中的日期
    print("\n正在识别排班表日期...")
    date_info = find_dates_in_schedule(ws)
    print(f"排班表日期: {list(date_info.keys())}")
    
    # 获取出勤表中的日期
    att_dates = list(att_date_objects.keys())
    print(f"出勤表日期: {att_dates}")
    
    if not date_info:
        print("错误：排班表中未找到日期！")
        messagebox.showerror("错误", "排班表中未找到日期！")
        return 0
    
    if not att_dates:
        print("错误：出勤表中未找到日期列！")
        messagebox.showerror("错误", "出勤表中未找到日期列！")
        return 0
    
    # 排序
    template_dates = list(date_info.keys())
    template_dates.sort()
    att_dates_sorted = att_dates.copy()
    att_dates_sorted.sort()
    
    print(f"\n出勤表日期(排序后): {att_dates_sorted}")
    print(f"排班表日期(排序后): {template_dates}")
    
    # 建立日期映射
    date_mapping = {}
    for i, template_date in enumerate(template_dates):
        if i < len(att_dates_sorted):
            date_mapping[template_date] = att_dates_sorted[i]
            print(f"  映射: {template_date} -> {att_dates_sorted[i]}")
    
    total_filled = 0
    
    # 为每个日期填写
    for template_date, (work_col, over_col) in date_info.items():
        if template_date in date_mapping:
            att_key = date_mapping[template_date]
            
            # 获取该日出勤的人员（从出勤表）
            if att_key in attendance_by_date:
                attendants = attendance_by_date[att_key]
                print(f"\n{template_date}: 出勤表有 {len(attendants)} 人出勤")
            else:
                attendants = {}
                print(f"\n{template_date}: 出勤表中无此日期数据")
            
            # 合并人员名单：排班表名单 + 出勤表出勤人员（去重）
            all_staff = set(template_staff) | set(attendants.keys())
            all_staff = sorted(list(all_staff))  # 排序
            
            print(f"合并后共 {len(all_staff)} 人需要处理")
            
            # 从起始行开始连续填写
            row_idx = 0
            for staff_name in all_staff:
                # 检查该员工是否出勤
                if staff_name in attendants:
                    overtime = attendants[staff_name]
                    current_row = start_row + row_idx
                    
                    # 填写姓名
                    ws.cell(current_row, work_col).value = staff_name
                    ws.cell(current_row, work_col).alignment = Alignment(horizontal='left')
                    
                    # 填写加班标记
                    mark = get_overtime_mark(overtime)
                    if mark:
                        ws.cell(current_row, over_col).value = mark
                        ws.cell(current_row, over_col).alignment = Alignment(horizontal='center')
                    
                    total_filled += 1
                    row_idx += 1
                else:
                    # 在排班表名单中但不出勤：不填写，不占位
                    pass
        else:
            print(f"跳过 {template_date}: 无对应出勤表日期")
    
    print(f"\n共填写 {total_filled} 人次")
    
    wb.save(output_path)
    wb.close()
    return total_filled

def main():
    att_file = select_file("请选择出勤表（第一张表）", "attendance")
    if not att_file:
        messagebox.showerror("错误", "未选择出勤表文件")
        return
    
    template_file = select_file("请选择排班表模板（第二张表）", "template")
    if not template_file:
        messagebox.showerror("错误", "未选择排班表模板")
        return
    
    try:
        # 读取第一张表（出勤表）
        df_att, att_date_map, att_date_objects, name_col, overtime_col, attendance_by_date = read_attendance(att_file)
        if df_att is None:
            messagebox.showerror("错误", "无法读取出勤表文件")
            return
        
        print(f"\n出勤表日期列: {list(att_date_map.keys())}")
        
        # 生成输出文件
        output_file = os.path.splitext(template_file)[0] + "_填写完成.xlsm"
        
        # 填写第二张表（排班表）
        total = fill_schedule(attendance_by_date, att_date_map, att_date_objects, template_file, output_file)
        
        if total > 0:
            messagebox.showinfo("完成", f"可出勤人员表填写完成！\n共标记 {total} 人次\n\n填写逻辑：\n- 排班表原有名单中的出勤人员\n- 出勤表中出勤但不在名单中的人员")
        else:
            messagebox.showwarning("警告", "没有标记任何人次！")
        
    except Exception as e:
        messagebox.showerror("错误", f"填写失败：{str(e)}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    main()