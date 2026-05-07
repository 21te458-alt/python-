import csv
import re
from datetime import datetime
import tkinter as tk
from tkinter import filedialog, messagebox
import os
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment, Font, PatternFill

# 关键词匹配规则
KEYWORD_RULES = {
    "提交时间": ["提交时间", "時間戳記", "timestamp", "time"],
    "姓名": ["姓名", "name", "名前", "氏名", "Name"],
    "加班情况": ["加班", "overtime", "残業", "availability", "Overtime"],
    "通行证": ["通行证", "priority pass", "pass", "カード", "Priority"],
    "签证类型": ["签证", "visa", "ビザ", "在留資格", "Visa"]
}

# 翻译映射表
TRANSLATION = {
    # 加班情况
    "A.30分钟以内": "A.30分钟以内",
    "B.30分钟到1小时": "B.30分钟到1小时",
    "C.1小时以上": "C.1小时以上",
    "D.不能加班": "D.不能加班",
    "A. Able to work overtime for less than 30 minutes": "A.30分钟以内",
    "B. Able to work overtime for 30 minutes to 1 hour": "B.30分钟到1小时",
    "C. Able to work overtime for more than 1 hour": "C.1小时以上",
    "D. Unable to work overtime": "D.不能加班",
    "Able to work overtime for less than 30 minutes": "A.30分钟以内",
    "Able to work overtime for 30 minutes to 1 hour": "B.30分钟到1小时",
    "Able to work overtime for more than 1 hour": "C.1小时以上",
    "Unable to work overtime": "D.不能加班",
    
    # 通行证
    "有": "有",
    "Yes": "有",
    "Maybe": "不确定",
    "No": "无",
    
    # 签证类型
    "留学签证": "留学签证",
    "家属滞在": "家属滞在",
    "其他": "其他",
    "日本人配偶/永住者配偶": "日本人配偶/永住者配偶",
    "Dependent visa (family members)": "家属滞在",
    "Dependent visa": "家属滞在",
    "Student visa": "留学签证",
    "Working visa": "工作签证",
    "Permanent resident": "永住者",
    "Spouse of Japanese citizen": "日本人配偶",
}

DATE_PATTERN = re.compile(r'(\d{1,2})[/\-月](\d{1,2})')

def translate_value(value):
    """翻译字段值"""
    if not value or value == '':
        return value
    # 精确匹配
    if value in TRANSLATION:
        return TRANSLATION[value]
    # 模糊匹配（去掉前后空格）
    stripped = value.strip()
    if stripped in TRANSLATION:
        return TRANSLATION[stripped]
    # 部分匹配（用于较长字符串）
    for key, trans in TRANSLATION.items():
        if key.lower() in stripped.lower() or stripped.lower() in key.lower():
            return trans
    return value

def is_date_column(col_name):
    col_str = str(col_name)
    skip_words = ['时间', '姓名', 'name', '签证', 'visa', '加班', 'overtime', '通行证', 'pass', 'timestamp', '提交者']
    col_lower = col_str.lower()
    for word in skip_words:
        if word in col_lower:
            return False
    
    match = DATE_PATTERN.search(col_str)
    if match:
        month, day = int(match.group(1)), int(match.group(2))
        if 1 <= month <= 12 and 1 <= day <= 31:
            return True
    return False

def parse_date_from_column(col_name):
    col_str = str(col_name)
    match = DATE_PATTERN.search(col_str)
    if match:
        return int(match.group(1)), int(match.group(2))
    return None

def read_excel_file(file_path):
    """使用 openpyxl 读取 Excel"""
    wb = load_workbook(file_path, data_only=True)
    ws = wb.active
    
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return None, None
    
    headers = [str(cell) if cell is not None else '' for cell in rows[0]]
    data_rows = []
    for row in rows[1:]:
        data_rows.append([str(cell) if cell is not None else '' for cell in row])
    
    return headers, data_rows

def read_csv_file(file_path):
    rows = []
    with open(file_path, 'r', encoding='utf-8-sig') as f:
        reader = csv.reader(f)
        headers = next(reader)
        for row in reader:
            rows.append(row)
    return headers, rows

def read_file(file_path):
    if file_path.endswith('.csv'):
        return read_csv_file(file_path)
    elif file_path.endswith('.xlsx'):
        return read_excel_file(file_path)
    else:
        return None, None

def find_columns_by_keywords(headers, rules):
    mapping = {}
    for target, keywords in rules.items():
        for i, col in enumerate(headers):
            col_lower = str(col).lower()
            if any(kw.lower() in col_lower for kw in keywords):
                mapping[target] = i
                break
    return mapping

def normalize_attendance(value):
    v = str(value).strip()
    if '出勤' in v or '✔' in v or v == '出':
        return '出勤'
    if '休息' in v or '休み' in v or '✖' in v or v == '休':
        return '休息'
    return v

def load_and_extract(file_path):
    print(f"\n正在处理: {os.path.basename(file_path)}")
    
    headers, rows = read_file(file_path)
    if headers is None:
        print(f"  错误: 无法读取文件")
        return None
    
    print(f"  原始表头: {headers[:5]}...")
    
    keyword_mapping = find_columns_by_keywords(headers, KEYWORD_RULES)
    print(f"  关键词匹配: {keyword_mapping}")
    
    date_indices = []
    for i, col in enumerate(headers):
        if is_date_column(col):
            date_indices.append((i, col))
    
    print(f"  日期列数: {len(date_indices)}")
    
    # 按日期排序
    date_indices_with_parsed = []
    for idx, col in date_indices:
        parsed = parse_date_from_column(col)
        if parsed:
            month, day = parsed
            try:
                date_obj = datetime(2026, month, day)
                date_indices_with_parsed.append((date_obj, idx, col))
            except:
                date_indices_with_parsed.append((None, idx, col))
        else:
            date_indices_with_parsed.append((None, idx, col))
    
    date_indices_with_parsed.sort(key=lambda x: (x[0] is None, x[0]))
    
    result_rows = []
    
    for row in rows:
        if len(row) < len(headers):
            row.extend([''] * (len(headers) - len(row)))
        
        result_row = {}
        
        # 提交时间
        if '提交时间' in keyword_mapping:
            idx = keyword_mapping['提交时间']
            result_row['提交时间'] = row[idx] if idx < len(row) else ''
        else:
            result_row['提交时间'] = ''
        
        # 姓名
        if '姓名' in keyword_mapping:
            idx = keyword_mapping['姓名']
            result_row['姓名'] = row[idx] if idx < len(row) else ''
        else:
            result_row['姓名'] = ''
        
        # 日期列
        for date_obj, idx, col in date_indices_with_parsed:
            if date_obj:
                col_name = f"{date_obj.month}月{date_obj.day}日"
            else:
                col_name = str(col)
            value = row[idx] if idx < len(row) else ''
            result_row[col_name] = normalize_attendance(value)
        
        # 加班情况（需要翻译）
        if '加班情况' in keyword_mapping:
            idx = keyword_mapping['加班情况']
            value = row[idx] if idx < len(row) else ''
            result_row['加班情况统计'] = translate_value(value)
        else:
            result_row['加班情况统计'] = ''
        
        # 通行证（需要翻译）
        if '通行证' in keyword_mapping:
            idx = keyword_mapping['通行证']
            value = row[idx] if idx < len(row) else ''
            result_row['是否有长期通行证'] = translate_value(value)
        else:
            result_row['是否有长期通行证'] = ''
        
        # 签证类型（需要翻译）
        if '签证类型' in keyword_mapping:
            idx = keyword_mapping['签证类型']
            value = row[idx] if idx < len(row) else ''
            result_row['签证类型'] = translate_value(value)
        else:
            result_row['签证类型'] = ''
        
        result_rows.append(result_row)
    
    print(f"  提取完成，共 {len(result_rows)} 行")
    if result_rows:
        print(f"  示例数据: 姓名={result_rows[0].get('姓名', '')}, 加班={result_rows[0].get('加班情况统计', '')}")
    
    return result_rows

def save_to_excel(result_rows):
    """保存为 Excel 文件，文件名使用当前日期"""
    if not result_rows:
        print("没有数据可保存！")
        return None
    
    current_date = datetime.now().strftime("%Y%m%d")
    filename = f"{current_date}_合并表格.xlsx"
    file_path = os.path.join(os.getcwd(), filename)
    
    # 创建 Excel 工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = "加班统计"
    
    # 获取所有列名
    all_keys = list(result_rows[0].keys())
    
    # 表头样式
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    
    # 写入表头
    for col_idx, key in enumerate(all_keys, 1):
        cell = ws.cell(row=1, column=col_idx, value=key)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # 写入数据
    for row_idx, row_data in enumerate(result_rows, 2):
        for col_idx, key in enumerate(all_keys, 1):
            value = row_data.get(key, '')
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # 自动调整列宽
    for col_idx, key in enumerate(all_keys, 1):
        max_length = len(key)
        for row_idx in range(2, min(len(result_rows) + 2, 100)):
            value = str(ws.cell(row=row_idx, column=col_idx).value)
            max_length = max(max_length, len(value))
        adjusted_width = min(max_length + 2, 25)
        col_letter = chr(64 + col_idx) if col_idx <= 26 else chr(64 + (col_idx - 1) // 26) + chr(64 + (col_idx - 1) % 26 + 1)
        ws.column_dimensions[col_letter].width = adjusted_width
    
    # 冻结首行
    ws.freeze_panes = 'A2'
    
    wb.save(file_path)
    return file_path

def select_files():
    root = tk.Tk()
    root.withdraw()
    root.attributes('-topmost', True)
    
    file1 = filedialog.askopenfilename(
        title="请选择第一个表格文件（中文表）",
        filetypes=[("Excel files", "*.xlsx"), ("CSV files", "*.csv"), ("All files", "*.*")]
    )
    
    if not file1:
        return None, None
    
    file2 = filedialog.askopenfilename(
        title="请选择第二个表格文件（英文/日文表）",
        filetypes=[("Excel files", "*.xlsx"), ("CSV files", "*.csv"), ("All files", "*.*")]
    )
    
    if not file2:
        return None, None
    
    root.destroy()
    return file1, file2

def main():
    print("=" * 50)
    print("表格合并工具 - 自动翻译版")
    print("=" * 50)
    
    file1, file2 = select_files()
    if not file1 or not file2:
        print("未选择文件，退出")
        return
    
    print(f"\n文件1: {os.path.basename(file1)}")
    print(f"文件2: {os.path.basename(file2)}")
    
    data1 = load_and_extract(file1)
    data2 = load_and_extract(file2)
    
    if data1 is None or data2 is None:
        print("\n文件读取失败！")
        messagebox.showerror("错误", "文件读取失败")
        return
    
    print(f"\n数据1: {len(data1)} 行")
    print(f"数据2: {len(data2)} 行")
    
    merged = data1 + data2
    print(f"合并后: {len(merged)} 行")
    
    if len(merged) == 0:
        messagebox.showwarning("警告", "合并后没有数据")
        return
    
    merged.sort(key=lambda x: x.get('提交时间', ''))
    
    save_path = save_to_excel(merged)
    
    if save_path:
        print(f"\n✅ 已保存: {save_path}")
        print(f"   共 {len(merged)} 行，{len(merged[0].keys())} 列")
        
        # 预览翻译后的数据
        print("\n翻译后预览（前3行）：")
        for i, row in enumerate(merged[:3]):
            name = row.get('姓名', '未知')
            overtime = row.get('加班情况统计', '')
            visa = row.get('签证类型', '')
            print(f"  {i+1}. {name} | 加班: {overtime} | 签证: {visa}")
        
        root = tk.Tk()
        root.withdraw()
        result = messagebox.askyesno("完成", f"合并完成！\n共 {len(merged)} 条记录\n\n文件：{os.path.basename(save_path)}\n\n是否打开文件夹？")
        root.destroy()
        
        if result:
            os.startfile(os.path.dirname(save_path))
    else:
        messagebox.showerror("错误", "保存失败")

if __name__ == "__main__":
    main()